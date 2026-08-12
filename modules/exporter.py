"""Exportación de licitaciones a CSV y Excel listos para descargar."""

from __future__ import annotations

import io
from datetime import datetime
from typing import Sequence

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.ingestion import COLUMN_LABELS

#: Orden de columnas preferido en las exportaciones.
EXPORT_COLUMNS: tuple[str, ...] = (
    "relevancia",
    "categoria",
    "badge",
    "expediente",
    "titulo",
    "organo_contratacion",
    "comunidad_autonoma",
    "fuente",
    "presupuesto_sin_iva",
    "presupuesto_con_iva",
    "ubicacion",
    "cpvs_texto",
    "cpvs_match",
    "keywords_match",
    "estado",
    "tipo_contrato",
    "fecha_actualizacion",
    "fecha_limite",
    "justificacion",
    "url",
)

CATEGORY_FILLS: dict[str, str] = {
    "Alta": "C8E6C9",
    "Media": "FFF3C4",
    "Baja": "F1F2F4",
}

HEADER_FILL = "1B4332"


def _stringify(valor: object) -> object:
    if isinstance(valor, (list, tuple, set)):
        return ", ".join(str(elemento) for elemento in valor)
    return valor


def prepare_export_dataframe(
    df: pd.DataFrame,
    columnas: Sequence[str] | None = None,
    usar_etiquetas: bool = True,
) -> pd.DataFrame:
    """Selecciona, ordena, aplana y renombra las columnas para la exportación."""
    seleccion = [c for c in (columnas or EXPORT_COLUMNS) if c in df.columns]
    if not seleccion:
        seleccion = list(df.columns)

    export = df[seleccion].copy()
    for columna in export.columns:
        export[columna] = export[columna].map(_stringify)

    if "fecha_actualizacion" in export.columns:
        export["fecha_actualizacion"] = pd.to_datetime(
            export["fecha_actualizacion"], errors="coerce"
        ).dt.strftime("%d/%m/%Y %H:%M")

    if usar_etiquetas:
        export = export.rename(columns={c: COLUMN_LABELS.get(c, c) for c in export.columns})
    return export.fillna("")


def timestamped_filename(prefijo: str = "licitaciones_grefa", extension: str = "csv") -> str:
    return f"{prefijo}_{datetime.now():%Y%m%d_%H%M}.{extension}"


def to_csv_bytes(df: pd.DataFrame, columnas: Sequence[str] | None = None) -> bytes:
    """CSV en UTF-8 con BOM y separador ';' (abre correctamente en Excel español)."""
    export = prepare_export_dataframe(df, columnas)
    buffer = io.StringIO()
    export.to_csv(buffer, index=False, sep=";", decimal=",")
    return buffer.getvalue().encode("utf-8-sig")


def to_excel_bytes(
    df: pd.DataFrame,
    columnas: Sequence[str] | None = None,
    nombre_hoja: str = "Licitaciones GREFA",
) -> bytes:
    """Excel con cabecera formateada, autofiltro, colores por categoría e hipervínculos."""
    export = prepare_export_dataframe(df, columnas)
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name=nombre_hoja[:31])
        hoja = writer.sheets[nombre_hoja[:31]]

        cabecera_fill = PatternFill("solid", fgColor=HEADER_FILL)
        cabecera_font = Font(color="FFFFFF", bold=True)
        for celda in hoja[1]:
            celda.fill = cabecera_fill
            celda.font = cabecera_font
            celda.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        hoja.freeze_panes = "A2"
        if hoja.max_row >= 1:
            hoja.auto_filter.ref = hoja.dimensions

        encabezados = [celda.value for celda in hoja[1]]
        indice_categoria = (
            encabezados.index(COLUMN_LABELS["categoria"]) + 1
            if COLUMN_LABELS["categoria"] in encabezados
            else None
        )
        indice_url = (
            encabezados.index(COLUMN_LABELS["url"]) + 1 if COLUMN_LABELS["url"] in encabezados else None
        )
        indices_importe = [
            encabezados.index(etiqueta) + 1
            for clave in ("presupuesto_sin_iva", "presupuesto_con_iva")
            if (etiqueta := COLUMN_LABELS[clave]) in encabezados
        ]

        for fila in range(2, hoja.max_row + 1):
            if indice_categoria:
                categoria = hoja.cell(row=fila, column=indice_categoria).value
                color = CATEGORY_FILLS.get(str(categoria))
                if color:
                    relleno = PatternFill("solid", fgColor=color)
                    for columna in range(1, hoja.max_column + 1):
                        hoja.cell(row=fila, column=columna).fill = relleno
            if indice_url:
                celda = hoja.cell(row=fila, column=indice_url)
                if isinstance(celda.value, str) and celda.value.startswith("http"):
                    celda.hyperlink = celda.value
                    celda.value = "Ver en PLACSP"
                    celda.font = Font(color="0563C1", underline="single")
            for columna in indices_importe:
                hoja.cell(row=fila, column=columna).number_format = '#,##0.00 "€"'

        anchos = {
            COLUMN_LABELS["titulo"]: 60,
            COLUMN_LABELS["organo_contratacion"]: 38,
            COLUMN_LABELS["justificacion"]: 45,
            COLUMN_LABELS["descripcion"]: 50,
        }
        for indice, encabezado in enumerate(encabezados, start=1):
            ancho = anchos.get(encabezado)
            if ancho is None:
                valores = [len(str(hoja.cell(row=f, column=indice).value or "")) for f in range(1, min(hoja.max_row, 200) + 1)]
                ancho = min(max([len(str(encabezado))] + valores) + 2, 40)
            hoja.column_dimensions[get_column_letter(indice)].width = ancho

    buffer.seek(0)
    return buffer.getvalue()
